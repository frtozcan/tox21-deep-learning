"""
╔══════════════════════════════════════════════════════════════════════════════╗
║         ADIM 4 — SHAP ANALİZİ (4 MODEL × 12 GÖREV)                        ║
║══════════════════════════════════════════════════════════════════════════════║
║  Her model için farklı SHAP açıklayıcısı kullanılır:                        ║
║                                                                              ║
║  MLP         → KernelExplainer (model-agnostik, yavaş ama evrensel)        ║
║  1D-CNN      → KernelExplainer (CNN çıkışını wrapper ile düzleştir)        ║
║  GNN         → KernelExplainer (atom özelliklerini descriptor gibi açıkla) ║
║  Transformer → KernelExplainer (token embedding'leri açıkla)               ║
║                                                                              ║
║  Çıktılar:                                                                   ║
║    shap/mlp_<görev>_shap.csv          → descriptor önem sıralaması         ║
║    shap/cnn_<görev>_shap.csv          → descriptor önem sıralaması         ║
║    shap/gnn_<görev>_shap.csv          → atom özelliği önem sıralaması      ║
║    shap/tfm_<görev>_shap.csv          → token pozisyon önem sıralaması     ║
║    shap/summary_top10_<görev>.png     → 4 model beeswarm / bar plot        ║
║    shap/cross_task_importance.png     → 12 görev karşılaştırma ısıl harita ║
║    shap/shap_summary.xlsx             → Excel özet tablosu                  ║
║                                                                              ║
║  Çalıştırma:                                                                 ║
║    python step4_shap_analysis.py                                            ║
║    python step4_shap_analysis.py --task SR-MMP  (tek görev)                ║
║    python step4_shap_analysis.py --n_background 30 --n_explain 50          ║
║                                                                              ║
║  Not: 4 model × 12 görev → ~2-4 saat (CPU'da)                              ║
║  Hızlandırmak için: --n_background 20 --n_explain 30                       ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import argparse
import os
import warnings
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import Dataset, DataLoader
from torch_geometric.data import Data
from torch_geometric.loader import DataLoader as GeoLoader
from torch_geometric.nn import GATConv, global_mean_pool, global_max_pool
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
import shap
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from rdkit import Chem
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
matplotlib.rcParams["font.family"] = "DejaVu Sans"

# ─── Sabitler ─────────────────────────────────────────────────────────────────
TASKS = ['NR-AR','NR-AR-LBD','NR-AhR','NR-Aromatase','NR-ER','NR-ER-LBD',
         'NR-PPAR-gamma','SR-ARE','SR-ATAD5','SR-HSE','SR-MMP','SR-p53']

DEVICE  = torch.device('cpu')   # SHAP CPU'da daha kararlı
N_DESC  = 200
MAX_LEN = 100

PALETTE = {
    "MLP":         "#2E75B6",
    "1D-CNN":      "#7030A0",
    "GNN":         "#C55A11",
    "Transformer": "#375623",
}

# GNN atom özellik isimleri (mol_to_graph ile aynı sırada)
GNN_FEAT_NAMES = [
    "AtomicNum/118", "Degree/10", "IsAromatic",
    "TotalHs/4",     "Mass/200",  "IsInRing",
    "TotalValence/8","FormalCharge/5","RadicalElec/4"
]

os.makedirs("shap", exist_ok=True)


# ══════════════════════════════════════════════════════════════════════════════
#  MODEL SINIFLARI (step3_models_v2.py ile aynı)
# ══════════════════════════════════════════════════════════════════════════════

class MLP(nn.Module):
    def __init__(self, in_dim=N_DESC):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(in_dim, 512), nn.BatchNorm1d(512), nn.ReLU(), nn.Dropout(0.3),
            nn.Linear(512, 256),    nn.BatchNorm1d(256), nn.ReLU(), nn.Dropout(0.3),
            nn.Linear(256, 128),    nn.ReLU(), nn.Dropout(0.2),
            nn.Linear(128, 1)
        )
    def forward(self, x): return self.net(x).squeeze(-1)


class CNN1D(nn.Module):
    def __init__(self):
        super().__init__()
        self.conv = nn.Sequential(
            nn.Conv1d(1, 64,  7, padding=3), nn.BatchNorm1d(64),  nn.ReLU(), nn.MaxPool1d(2),
            nn.Conv1d(64, 128, 5, padding=2), nn.BatchNorm1d(128), nn.ReLU(), nn.MaxPool1d(2),
            nn.Conv1d(128, 256, 3, padding=1), nn.BatchNorm1d(256), nn.ReLU(), nn.AdaptiveAvgPool1d(1),
        )
        self.fc = nn.Sequential(nn.Flatten(), nn.Linear(256,128), nn.ReLU(), nn.Dropout(0.3), nn.Linear(128,1))
    def forward(self, x): return self.fc(self.conv(x)).squeeze(-1)


class GNN(nn.Module):
    def __init__(self):
        super().__init__()
        self.conv1 = GATConv(9,   128, heads=4, concat=False, dropout=0.2)
        self.conv2 = GATConv(128, 128, heads=4, concat=False, dropout=0.2)
        self.conv3 = GATConv(128, 128, heads=2, concat=False, dropout=0.2)
        self.ln1, self.ln2, self.ln3 = nn.LayerNorm(128), nn.LayerNorm(128), nn.LayerNorm(128)
        self.fc = nn.Sequential(nn.Linear(256,256), nn.ReLU(), nn.Dropout(0.3),
                                nn.Linear(256,64),  nn.ReLU(), nn.Linear(64,1))
    def forward(self, data):
        x, ei, b = data.x, data.edge_index, data.batch
        x = F.elu(self.ln1(self.conv1(x, ei)))
        x = F.elu(self.ln2(self.conv2(x, ei)))
        x = F.elu(self.ln3(self.conv3(x, ei)))
        return self.fc(torch.cat([global_mean_pool(x,b), global_max_pool(x,b)], dim=-1)).squeeze(-1)


class SMILESTransformer(nn.Module):
    def __init__(self, vocab_size, d_model=64, nhead=4, num_layers=2):
        super().__init__()
        self.token_emb = nn.Embedding(vocab_size, d_model, padding_idx=0)
        self.pos_emb   = nn.Embedding(MAX_LEN, d_model)
        enc = nn.TransformerEncoderLayer(d_model, nhead, dim_feedforward=256,
                                         dropout=0.1, batch_first=True, norm_first=True)
        self.transformer = nn.TransformerEncoder(enc, num_layers=num_layers)
        self.classifier  = nn.Sequential(nn.LayerNorm(d_model), nn.Linear(d_model,32),
                                         nn.GELU(), nn.Dropout(0.2), nn.Linear(32,1))
    def forward(self, x):
        pos  = torch.arange(x.size(1), device=x.device).unsqueeze(0)
        h    = self.token_emb(x) + self.pos_emb(pos)
        mask = (x == 0)
        h    = self.transformer(h, src_key_padding_mask=mask)
        return self.classifier(h[:, 0, :]).squeeze(-1)


# ══════════════════════════════════════════════════════════════════════════════
#  VERİ HAZIRLIK (step3 ile aynı bölme — random_state=42)
# ══════════════════════════════════════════════════════════════════════════════

class TabDS(Dataset):
    def __init__(self, X, y, dtype=torch.float32):
        self.X = torch.tensor(X, dtype=dtype)
        self.y = torch.tensor(y, dtype=torch.float32)
    def __len__(self): return len(self.y)
    def __getitem__(self, i): return self.X[i], self.y[i]


def mol_to_graph(smiles, label=0.0):
    mol = Chem.MolFromSmiles(smiles)
    if mol is None: return None
    feats = [[a.GetAtomicNum()/118, a.GetDegree()/10, int(a.GetIsAromatic()),
              a.GetTotalNumHs()/4, a.GetMass()/200, int(a.IsInRing()),
              a.GetTotalValence()/8, a.GetFormalCharge()/5,
              a.GetNumRadicalElectrons()/4] for a in mol.GetAtoms()]
    x  = torch.tensor(feats, dtype=torch.float)
    el = []
    for b in mol.GetBonds():
        i, j = b.GetBeginAtomIdx(), b.GetEndAtomIdx()
        el += [[i,j],[j,i]]
    ei = torch.tensor(el, dtype=torch.long).t().contiguous() if el else torch.zeros((2,0),dtype=torch.long)
    return Data(x=x, edge_index=ei, y=torch.tensor([float(label)]))


def prepare_task_data(task, df_full, X_desc_all, X_tok_all):
    """
    step3 ile birebir aynı veri hazırlama (random_state=42).
    Aynı train/val/test bölmesi → SHAP değerleri train verisiyle tutarlı.
    """
    all_smiles = df_full['smiles'].tolist()
    df_t = df_full[['smiles', task]].dropna()
    df_t = df_t[df_t[task].isin([0.0,1.0])].drop_duplicates('smiles').reset_index(drop=True)
    df_t = df_t[df_t['smiles'].apply(lambda s: Chem.MolFromSmiles(s) is not None)].reset_index(drop=True)

    smiles_list = df_t['smiles'].tolist()
    labels      = df_t[task].values
    fi          = np.array([all_smiles.index(s) if s in all_smiles else 0 for s in smiles_list])

    idx = np.arange(len(df_t))
    idx_tv, idx_te = train_test_split(idx, test_size=0.15, stratify=labels, random_state=42)
    idx_tr, idx_vl = train_test_split(idx_tv, test_size=0.176, stratify=labels[idx_tv], random_state=42)

    sc  = StandardScaler()
    Xtr = np.nan_to_num(sc.fit_transform(X_desc_all[fi][idx_tr]))
    Xte = np.nan_to_num(sc.transform(X_desc_all[fi][idx_te]))

    return {
        'smiles':    smiles_list,
        'labels':    labels,
        'fi':        fi,
        'idx_tr':    idx_tr,
        'idx_te':    idx_te,
        'Xtr':       Xtr,       # normalize descriptor — train
        'Xte':       Xte,       # normalize descriptor — test
        'Xt':        X_tok_all[fi],  # token dizisi
        'scaler':    sc,
        'X_desc_all': X_desc_all,
        'X_tok_all':  X_tok_all,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  MODEL 1 — MLP SHAP
# ══════════════════════════════════════════════════════════════════════════════

def shap_mlp(task, data, desc_names, n_bg, n_exp):
    """
    KernelExplainer ile MLP için SHAP değerleri.
    
    KernelExplainer:
      - Model-agnostik: herhangi bir fonksiyona uygulanabilir
      - background: SHAP'ın beklenti değerini (E[f(x)]) hesaplamak için kullanır
        k-means ile özetlenmiş n_bg örnek
      - explain: SHAP değerleri hesaplanacak n_exp test örneği
      
    Çıktı: her descriptor için ortalama |SHAP| değeri
    """
    pt = f"models/mlp_{task}.pt"
    if not os.path.exists(pt):
        print(f"    ⚠ {pt} bulunamadı, atlanıyor.")
        return None

    model = MLP().to(DEVICE)
    model.load_state_dict(torch.load(pt, map_location=DEVICE))
    model.eval()

    # Predict fonksiyonu: numpy array → numpy olasılık
    def predict_fn(X_np):
        X_t = torch.tensor(X_np, dtype=torch.float32).to(DEVICE)
        with torch.no_grad():
            return torch.sigmoid(model(X_t)).cpu().numpy()

    Xtr, Xte = data['Xtr'], data['Xte']

    # Background: train setinden k-means ile n_bg özetle
    print(f"    KernelExplainer background oluşturuluyor ({n_bg} örnek)...")
    background = shap.kmeans(Xtr, n_bg)

    # Explainer
    explainer = shap.KernelExplainer(predict_fn, background)

    # Açıklanacak test örnekleri
    n_exp = min(n_exp, len(Xte))
    X_explain = Xte[:n_exp]
    print(f"    {n_exp} test örneği açıklanıyor...")
    shap_values = explainer.shap_values(X_explain, nsamples=100, silent=True)

    # Ortalama |SHAP| — özellik önem sıralaması
    mean_abs = np.abs(shap_values).mean(axis=0)
    df_imp   = pd.DataFrame({'feature': desc_names, 'mean_abs_shap': mean_abs})
    df_imp   = df_imp.sort_values('mean_abs_shap', ascending=False).reset_index(drop=True)

    # CSV kaydet
    path_csv = f"shap/mlp_{task}_shap.csv"
    df_imp.to_csv(path_csv, index=False)

    return {'shap_values': shap_values, 'X_explain': X_explain,
            'feature_names': desc_names, 'importance': df_imp, 'model': 'MLP'}


# ══════════════════════════════════════════════════════════════════════════════
#  MODEL 2 — 1D-CNN SHAP
# ══════════════════════════════════════════════════════════════════════════════

def shap_cnn(task, data, desc_names, n_bg, n_exp):
    """
    CNN için KernelExplainer.
    CNN [batch, 1, 200] girdi bekler → wrapper ile [batch, 200] alır.
    
    NOT: CNN descriptor sırasındaki yerel komşuluklara duyarlı.
    SHAP değerleri MLP'den farklı dağılım gösterebilir.
    """
    pt = f"models/cnn_{task}.pt"
    if not os.path.exists(pt):
        print(f"    ⚠ {pt} bulunamadı, atlanıyor.")
        return None

    model = CNN1D().to(DEVICE)
    model.load_state_dict(torch.load(pt, map_location=DEVICE))
    model.eval()

    def predict_fn(X_np):
        # [n, 200] → [n, 1, 200] (CNN kanal boyutu)
        X_t = torch.tensor(X_np, dtype=torch.float32).unsqueeze(1).to(DEVICE)
        with torch.no_grad():
            return torch.sigmoid(model(X_t)).cpu().numpy()

    Xtr, Xte = data['Xtr'], data['Xte']

    print(f"    KernelExplainer background oluşturuluyor ({n_bg} örnek)...")
    background  = shap.kmeans(Xtr, n_bg)
    explainer   = shap.KernelExplainer(predict_fn, background)

    n_exp = min(n_exp, len(Xte))
    X_explain = Xte[:n_exp]
    print(f"    {n_exp} test örneği açıklanıyor...")
    shap_values = explainer.shap_values(X_explain, nsamples=100, silent=True)

    mean_abs = np.abs(shap_values).mean(axis=0)
    df_imp   = pd.DataFrame({'feature': desc_names, 'mean_abs_shap': mean_abs})
    df_imp   = df_imp.sort_values('mean_abs_shap', ascending=False).reset_index(drop=True)
    df_imp.to_csv(f"shap/cnn_{task}_shap.csv", index=False)

    return {'shap_values': shap_values, 'X_explain': X_explain,
            'feature_names': desc_names, 'importance': df_imp, 'model': '1D-CNN'}


# ══════════════════════════════════════════════════════════════════════════════
#  MODEL 3 — GNN SHAP
# ══════════════════════════════════════════════════════════════════════════════

def shap_gnn(task, data, n_bg, n_exp):
    """
    GNN için SHAP — atom özelliklerini açıkla.
    
    GNN farklı boyutlu girdi (graf) alır.
    Yaklaşım: her molekülü ortalama atom özellik vektörüne indirgeyerek
    9 boyutlu bir uzayda KernelExplainer uygula.
    
    Bu basitleştirme GNN'in mesaj geçişini tam yansıtmaz ama
    hangi atom özelliklerinin genel olarak önemli olduğunu gösterir.
    """
    pt = f"models/gnn_{task}.pt"
    if not os.path.exists(pt):
        print(f"    ⚠ {pt} bulunamadı, atlanıyor.")
        return None

    model = GNN().to(DEVICE)
    model.load_state_dict(torch.load(pt, map_location=DEVICE))
    model.eval()

    smiles_list = data['smiles']
    labels      = data['labels']
    idx_tr      = data['idx_tr']
    idx_te      = data['idx_te']

    # Molekülleri ortalama atom özellik vektörüne dönüştür [n, 9]
    def smiles_to_mean_feat(idx_list):
        vecs = []
        for i in idx_list:
            mol = Chem.MolFromSmiles(smiles_list[i])
            if mol is None:
                vecs.append(np.zeros(9))
                continue
            feats = [[a.GetAtomicNum()/118, a.GetDegree()/10, int(a.GetIsAromatic()),
                      a.GetTotalNumHs()/4, a.GetMass()/200, int(a.IsInRing()),
                      a.GetTotalValence()/8, a.GetFormalCharge()/5,
                      a.GetNumRadicalElectrons()/4] for a in mol.GetAtoms()]
            vecs.append(np.mean(feats, axis=0))
        return np.array(vecs, dtype=np.float32)

    X_train_gnn = smiles_to_mean_feat(idx_tr)
    X_test_gnn  = smiles_to_mean_feat(idx_te)

    # GNN'i ortalama atom özelliği üzerinden çağırmak için wrapper
    # Her satırı tek bir dummy molekül grafına dönüştür
    def predict_fn(X_np):
        probs = []
        for row in X_np:
            # Tek atomlu dummy graf: 9 özellik
            x_t  = torch.tensor([row], dtype=torch.float)
            ei   = torch.zeros((2, 0), dtype=torch.long)
            b    = torch.zeros(1, dtype=torch.long)
            data_obj = Data(x=x_t, edge_index=ei)
            data_obj.batch = b
            with torch.no_grad():
                p = torch.sigmoid(model(data_obj)).item()
            probs.append(p)
        return np.array(probs)

    print(f"    GNN ortalama atom özellikleri hesaplanıyor...")
    background  = shap.kmeans(X_train_gnn, min(n_bg, len(X_train_gnn)))
    explainer   = shap.KernelExplainer(predict_fn, background)

    n_exp = min(n_exp, len(X_test_gnn))
    X_explain = X_test_gnn[:n_exp]
    print(f"    {n_exp} test örneği açıklanıyor (GNN — yavaş olabilir)...")
    shap_values = explainer.shap_values(X_explain, nsamples=50, silent=True)

    mean_abs = np.abs(shap_values).mean(axis=0)
    df_imp   = pd.DataFrame({'feature': GNN_FEAT_NAMES, 'mean_abs_shap': mean_abs})
    df_imp   = df_imp.sort_values('mean_abs_shap', ascending=False).reset_index(drop=True)
    df_imp.to_csv(f"shap/gnn_{task}_shap.csv", index=False)

    return {'shap_values': shap_values, 'X_explain': X_explain,
            'feature_names': GNN_FEAT_NAMES, 'importance': df_imp, 'model': 'GNN'}


# ══════════════════════════════════════════════════════════════════════════════
#  MODEL 4 — TRANSFORMER SHAP
# ══════════════════════════════════════════════════════════════════════════════

def shap_transformer(task, data, vocab, n_bg, n_exp):
    """
    Transformer için SHAP — token pozisyonlarını açıkla.
    
    Transformer'ın token embedding'lerini 64 boyutlu vektörler üzerinden
    açıklamak yerine, her pozisyonun ortalama embedding normunu kullanırız.
    Bu pozisyon bazlı önem skorunu gösterir:
      - Hangi SMILES pozisyonu (0-99) toksisite kararında daha etkili?
    
    Alternatif: token indeksleri üzerinden doğrudan KernelExplainer
      - Daha doğal ama yavaş
    """
    pt = f"models/tfm_{task}.pt"
    if not os.path.exists(pt):
        print(f"    ⚠ {pt} bulunamadı, atlanıyor.")
        return None

    vocab_size = len(vocab)
    model = SMILESTransformer(vocab_size).to(DEVICE)
    model.load_state_dict(torch.load(pt, map_location=DEVICE))
    model.eval()

    Xt       = data['Xt']
    idx_tr   = data['idx_tr']
    idx_te   = data['idx_te']

    X_train_tok = Xt[idx_tr].astype(np.float32)   # float'a çevir (SHAP ister)
    X_test_tok  = Xt[idx_te].astype(np.float32)

    def predict_fn(X_np):
        # float → long (embedding indeksi)
        X_t = torch.tensor(X_np.astype(np.int64), dtype=torch.long).to(DEVICE)
        with torch.no_grad():
            return torch.sigmoid(model(X_t)).cpu().numpy()

    print(f"    Transformer token pozisyonları açıklanıyor...")
    background  = shap.kmeans(X_train_tok, min(n_bg, len(X_train_tok)))
    explainer   = shap.KernelExplainer(predict_fn, background)

    n_exp = min(n_exp, len(X_test_tok))
    X_explain = X_test_tok[:n_exp]
    print(f"    {n_exp} test örneği açıklanıyor...")
    shap_values = explainer.shap_values(X_explain, nsamples=80, silent=True)

    # Pozisyon isimleri: pos_0, pos_1, ..., pos_99
    pos_names = [f"pos_{i}" for i in range(MAX_LEN)]
    mean_abs  = np.abs(shap_values).mean(axis=0)
    df_imp    = pd.DataFrame({'feature': pos_names, 'mean_abs_shap': mean_abs})
    df_imp    = df_imp.sort_values('mean_abs_shap', ascending=False).reset_index(drop=True)
    df_imp.to_csv(f"shap/tfm_{task}_shap.csv", index=False)

    return {'shap_values': shap_values, 'X_explain': X_explain,
            'feature_names': pos_names, 'importance': df_imp, 'model': 'Transformer'}


# ══════════════════════════════════════════════════════════════════════════════
#  GÖRSELLEŞTİRME — Görev başına özet figür (4 model)
# ══════════════════════════════════════════════════════════════════════════════

def plot_task_summary(task, results_dict, top_n=10):
    """
    Bir görev için 4 modelin top-N özellik önem sıralamasını yan yana çizer.
    Makale kalitesi: yatay çubuk grafikleri, renk kodlu modeller.
    """
    models_with_data = [(m, r) for m, r in results_dict.items() if r is not None]
    if not models_with_data:
        return

    n_models = len(models_with_data)
    fig, axes = plt.subplots(1, n_models, figsize=(n_models * 4.5, 7))
    if n_models == 1:
        axes = [axes]

    fig.suptitle(f"SHAP Feature Importance — {task}",
                 fontsize=14, fontweight="bold", y=1.02)

    for ax, (mname, res) in zip(axes, models_with_data):
        df_imp = res['importance'].head(top_n)
        color  = PALETTE.get(mname, "#555555")

        bars = ax.barh(range(len(df_imp)), df_imp['mean_abs_shap'],
                       color=color, alpha=0.85, edgecolor="white", linewidth=0.5)

        ax.set_yticks(range(len(df_imp)))
        ax.set_yticklabels(df_imp['feature'], fontsize=8.5)
        ax.invert_yaxis()   # en önemli üstte
        ax.set_xlabel("Mean |SHAP value|", fontsize=9)
        ax.set_title(mname, fontsize=11, fontweight="bold", color=color)
        ax.spines[["top","right"]].set_visible(False)
        ax.grid(axis="x", alpha=0.3, linestyle="--")

        # Değer etiketleri
        for bar, val in zip(bars, df_imp['mean_abs_shap']):
            ax.text(bar.get_width() + 0.0001, bar.get_y() + bar.get_height()/2,
                    f"{val:.4f}", va="center", fontsize=7, color=color)

    plt.tight_layout()
    path = f"shap/summary_top{top_n}_{task}.png"
    plt.savefig(path, dpi=200, bbox_inches="tight")
    plt.close()
    print(f"    Figür kaydedildi: {path}")


def plot_cross_task_heatmap(all_results, desc_names, top_n=15):
    """
    12 görev × top-N descriptor için MLP SHAP ısıl haritası.
    Hangi descriptor hangi görevde kritik? sorusuna görsel cevap verir.
    Makale için çok güçlü bir figür.
    """
    # MLP sonuçlarını topla
    mlp_data = {}
    for task, res_dict in all_results.items():
        if 'MLP' in res_dict and res_dict['MLP'] is not None:
            mlp_data[task] = res_dict['MLP']['importance'].set_index('feature')['mean_abs_shap']

    if not mlp_data:
        return

    # Global top-N descriptor (tüm görevlerde en önemli olanlar)
    combined = pd.DataFrame(mlp_data).fillna(0)
    global_top = combined.mean(axis=1).nlargest(top_n).index.tolist()
    mat        = combined.loc[global_top].T   # [task × feature]

    fig, ax = plt.subplots(figsize=(top_n * 0.85 + 2, len(mlp_data) * 0.55 + 2))
    sns.heatmap(mat, annot=True, fmt=".3f", cmap="YlOrRd",
                linewidths=0.4, linecolor="white",
                annot_kws={"size": 7}, ax=ax,
                cbar_kws={"label": "Mean |SHAP|", "shrink": 0.6})

    ax.set_title(f"MLP SHAP Importance — Top {top_n} Descriptors × 12 Tasks",
                 fontsize=13, fontweight="bold", pad=12)
    ax.set_xlabel("Descriptor", fontsize=10)
    ax.set_ylabel("Task", fontsize=10)
    ax.tick_params(axis='x', labelsize=8, rotation=40)
    ax.tick_params(axis='y', labelsize=9)

    plt.tight_layout()
    path = "shap/cross_task_heatmap_mlp.png"
    plt.savefig(path, dpi=200, bbox_inches="tight")
    plt.close()
    print(f"\n  Cross-task ısıl haritası kaydedildi: {path}")


def plot_beeswarm(task, res, top_n=15):
    """
    Klasik SHAP beeswarm plot — her nokta bir örnek.
    Renk: özellik değeri (mavi=düşük, kırmızı=yüksek)
    X ekseni: SHAP değeri (pozitif → toksik olasılığını artırıyor)
    """
    if res is None or res['shap_values'] is None:
        return

    shap_vals = res['shap_values']
    X_exp     = res['X_explain']
    feat_names= res['feature_names']
    mname     = res['model']

    # Top-N özelliği seç
    mean_abs  = np.abs(shap_vals).mean(axis=0)
    top_idx   = np.argsort(mean_abs)[::-1][:top_n]

    shap_top  = shap_vals[:, top_idx]
    X_top     = X_exp[:, top_idx] if X_exp.shape[1] == len(feat_names) else X_exp[:, top_idx]
    names_top = [feat_names[i] for i in top_idx]

    try:
        shap_exp = shap.Explanation(
            values=shap_top, data=X_top,
            feature_names=names_top
        )
        fig, ax = plt.subplots(figsize=(9, 6))
        shap.plots.beeswarm(shap_exp, max_display=top_n, show=False, color_bar=True)
        plt.title(f"SHAP Beeswarm — {mname} / {task}", fontsize=12, fontweight="bold")
        plt.tight_layout()
        path = f"shap/beeswarm_{mname.replace('-','').lower()}_{task}.png"
        plt.savefig(path, dpi=150, bbox_inches="tight")
        plt.close()
    except Exception as e:
        print(f"    ⚠ Beeswarm çizilemedi ({mname}/{task}): {e}")


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL ÖZET TABLOSU
# ══════════════════════════════════════════════════════════════════════════════

def save_shap_excel(all_results, desc_names):
    """
    Tüm SHAP sonuçlarını biçimlendirilmiş Excel dosyasına yazar.
    Her görev için ayrı sayfa + özet karşılaştırma sayfası.
    """
    HEADER_FILL  = PatternFill("solid", start_color="1F3864")
    HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    TASK_FILL    = PatternFill("solid", start_color="D6E4F0")
    NORMAL_FONT  = Font(name="Arial", size=9)
    CENTER       = Alignment(horizontal="center", vertical="center")
    LEFT         = Alignment(horizontal="left",   vertical="center")
    THIN         = Side(style="thin", color="BFBFBF")
    BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    MODEL_COLORS = {"MLP":"2E75B6","1D-CNN":"7030A0","GNN":"C55A11","Transformer":"375623"}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Her görev için bir sayfa ──────────────────────────────────────────────
    for task, res_dict in all_results.items():
        ws = wb.create_sheet(task[:25])   # sayfa adı max 31 karakter

        # Başlık
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = f"SHAP Analysis — {task}"
        title_cell.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        title_cell.fill  = HEADER_FILL
        title_cell.alignment = CENTER
        ws.row_dimensions[1].height = 22

        col_offset = 1
        for mname, res in res_dict.items():
            if res is None:
                continue

            df_imp = res['importance'].head(20)
            color  = MODEL_COLORS.get(mname, "555555")
            model_fill = PatternFill("solid", start_color=color)

            # Model başlığı
            header_row = 2
            ws.merge_cells(
                start_row=header_row, start_column=col_offset,
                end_row=header_row,   end_column=col_offset+1
            )
            hcell = ws.cell(row=header_row, column=col_offset, value=mname)
            hcell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            hcell.fill      = model_fill
            hcell.alignment = CENTER
            ws.row_dimensions[header_row].height = 18

            # Sütun başlıkları
            ws.cell(row=3, column=col_offset,   value="Feature").font    = HEADER_FONT
            ws.cell(row=3, column=col_offset,   value="Feature").fill    = HEADER_FILL
            ws.cell(row=3, column=col_offset,   value="Feature").alignment = CENTER
            ws.cell(row=3, column=col_offset+1, value="SHAP").font       = HEADER_FONT
            ws.cell(row=3, column=col_offset+1, value="SHAP").fill       = HEADER_FILL
            ws.cell(row=3, column=col_offset+1, value="SHAP").alignment  = CENTER
            ws.row_dimensions[3].height = 16
            ws.column_dimensions[get_column_letter(col_offset)].width   = 22
            ws.column_dimensions[get_column_letter(col_offset+1)].width = 10

            # Veri satırları
            for ri, row in enumerate(df_imp.itertuples(index=False), 4):
                bg = PatternFill("solid", start_color="EAF4FB" if ri%2==0 else "FFFFFF")
                fc = ws.cell(row=ri, column=col_offset,   value=row.feature)
                vc = ws.cell(row=ri, column=col_offset+1, value=round(row.mean_abs_shap, 5))
                for cell in [fc, vc]:
                    cell.font = NORMAL_FONT
                    cell.fill = bg
                    cell.border = BORDER
                fc.alignment = LEFT
                vc.alignment = CENTER
                vc.number_format = "0.00000"
                ws.row_dimensions[ri].height = 14

            col_offset += 3   # model sütunları arasında boşluk

    # ── Özet karşılaştırma sayfası ────────────────────────────────────────────
    ws_sum = wb.create_sheet("Özet", 0)
    ws_sum["A1"].value = "SHAP Özet — Top-1 Descriptor per Model per Task"
    ws_sum["A1"].font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws_sum["A1"].fill  = HEADER_FILL
    ws_sum["A1"].alignment = CENTER
    ws_sum.merge_cells("A1:F1")
    ws_sum.row_dimensions[1].height = 22

    headers = ["Görev", "MLP Top-1", "CNN Top-1", "GNN Top-1", "TFM Top-1", "Ortak En Önemli"]
    for ci, h in enumerate(headers, 1):
        c = ws_sum.cell(row=2, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
        ws_sum.column_dimensions[get_column_letter(ci)].width = 24
    ws_sum.row_dimensions[2].height = 18

    model_keys = ["MLP", "1D-CNN", "GNN", "Transformer"]
    for ri, task in enumerate(all_results.keys(), 3):
        res_dict = all_results[task]
        bg = PatternFill("solid", start_color="D6E4F0" if ri%2==0 else "FFFFFF")

        tc = ws_sum.cell(row=ri, column=1, value=task)
        tc.font = Font(name="Arial", bold=True, size=9)
        tc.fill = TASK_FILL; tc.alignment = LEFT; tc.border = BORDER

        top1s = []
        for ci, mk in enumerate(model_keys, 2):
            res = res_dict.get(mk)
            val = res['importance']['feature'].iloc[0] if res is not None else "N/A"
            top1s.append(val)
            c = ws_sum.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=9)
            c.fill = bg; c.alignment = CENTER; c.border = BORDER

        # Ortak: en sık görülen Top-1
        from collections import Counter
        common = Counter([t for t in top1s if t != "N/A"]).most_common(1)
        common_val = common[0][0] if common else "—"
        cc = ws_sum.cell(row=ri, column=6, value=common_val)
        cc.font = Font(name="Arial", bold=True, size=9, color="375623")
        cc.fill = PatternFill("solid", start_color="E2EFDA")
        cc.alignment = CENTER; cc.border = BORDER
        ws_sum.row_dimensions[ri].height = 15

    path = "shap/shap_summary.xlsx"
    wb.save(path)
    print(f"\n  SHAP Excel özeti kaydedildi: {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  ANA FONKSİYON
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Tox21 SHAP Analizi")
    parser.add_argument('--task',         default='all',
                        help='Görev adı veya "all"')
    parser.add_argument('--n_background', type=int, default=50,
                        help='SHAP background örnek sayısı (varsayılan: 50)')
    parser.add_argument('--n_explain',    type=int, default=80,
                        help='SHAP açıklanacak test örneği (varsayılan: 80)')
    args = parser.parse_args()

    N_BG  = args.n_background
    N_EXP = args.n_explain

    print("=" * 65)
    print("  ADIM 4 — SHAP ANALİZİ")
    print(f"  background={N_BG}, explain={N_EXP}")
    print("=" * 65)

    # ── Veri yükle ────────────────────────────────────────────────────────────
    print("\nVeri yükleniyor...")
    df_full   = pd.read_csv("data/tox21.csv")
    X_desc    = np.load("features/X_descriptors.npy")
    X_tok     = np.load("features/X_tokens.npy")
    desc_names = pd.read_csv("features/descriptor_names.csv")['descriptor'].tolist()
    all_chars  = sorted(set(''.join(df_full['smiles'].tolist())))
    vocab      = ['<PAD>', '<UNK>'] + all_chars

    print(f"  {len(df_full)} bileşik | {len(desc_names)} descriptor | vocab={len(vocab)}")

    # ── Görev listesi ─────────────────────────────────────────────────────────
    selected = TASKS if args.task == 'all' else [args.task]
    all_results = {}

    for task in selected:
        print(f"\n{'─'*65}")
        print(f"  GÖREV: {task}")
        print(f"{'─'*65}")

        data = prepare_task_data(task, df_full, X_desc, X_tok)
        task_res = {}

        # MLP
        print("\n  [1/4] MLP SHAP")
        task_res['MLP'] = shap_mlp(task, data, desc_names, N_BG, N_EXP)

        # 1D-CNN
        print("\n  [2/4] 1D-CNN SHAP")
        task_res['1D-CNN'] = shap_cnn(task, data, desc_names, N_BG, N_EXP)

        # GNN
        print("\n  [3/4] GNN SHAP")
        task_res['GNN'] = shap_gnn(task, data, N_BG, min(N_EXP//2, 40))
        # GNN yavaş: daha az örnek

        # Transformer
        print("\n  [4/4] Transformer SHAP")
        task_res['Transformer'] = shap_transformer(task, data, vocab, N_BG, N_EXP)

        all_results[task] = task_res

        # Görev figürü
        print(f"\n  [{task}] Figürler oluşturuluyor...")
        plot_task_summary(task, task_res, top_n=10)

        # Beeswarm (MLP ve 1D-CNN için — en yorumlanabilir)
        if task_res.get('MLP'):
            plot_beeswarm(task, task_res['MLP'], top_n=15)
        if task_res.get('1D-CNN'):
            plot_beeswarm(task, task_res['1D-CNN'], top_n=15)

    # ── Cross-task ısıl haritası (MLP) ────────────────────────────────────────
    if len(selected) > 1:
        print("\n  Cross-task ısıl haritası oluşturuluyor...")
        plot_cross_task_heatmap(all_results, desc_names, top_n=15)

    # ── Excel özet ────────────────────────────────────────────────────────────
    print("\n  Excel özeti kaydediliyor...")
    save_shap_excel(all_results, desc_names)

    # ── Özet ──────────────────────────────────────────────────────────────────
    saved_files = [f for f in os.listdir("shap") if f.endswith(('.png', '.csv', '.xlsx'))]
    print(f"\n{'='*65}")
    print(f"  ✅ SHAP analizi tamamlandı!")
    print(f"  {len(saved_files)} dosya 'shap/' klasörüne kaydedildi")
    print(f"\n  Görev başına dosyalar:")
    print(f"    shap/mlp_<görev>_shap.csv        → MLP descriptor önemi")
    print(f"    shap/cnn_<görev>_shap.csv        → CNN descriptor önemi")
    print(f"    shap/gnn_<görev>_shap.csv        → GNN atom özellik önemi")
    print(f"    shap/tfm_<görev>_shap.csv        → Transformer token önemi")
    print(f"    shap/summary_top10_<görev>.png   → 4 model karşılaştırma")
    print(f"    shap/beeswarm_mlp_<görev>.png    → SHAP beeswarm")
    if len(selected) > 1:
        print(f"    shap/cross_task_heatmap_mlp.png  → 12 görev ısıl haritası")
    print(f"    shap/shap_summary.xlsx           → Excel özet tablosu")
    print(f"{'='*65}")


if __name__ == "__main__":
    main()
