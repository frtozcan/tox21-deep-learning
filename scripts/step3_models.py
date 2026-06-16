"""
ADIM 3 — 4 DERİN ÖĞRENME MODELİ
==================================
Model 1: MLP        — Descriptor tabanlı, tam bağlı sinir ağı
Model 2: 1D-CNN     — Descriptor dizisi üzerinde konvolüsyon
Model 3: GNN (GAT)  — Moleküler graf üzerinde attention mekanizması
Model 4: SMILES-TFM — Karakter dizisi üzerinde Transformer

Çalıştırma: python step3_models.py --task SR-MMP
            python step3_models.py --task all
"""

import argparse
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import Dataset, DataLoader
from torch_geometric.data import Data
from torch_geometric.loader import DataLoader as GeoLoader
from torch_geometric.nn import GATConv, global_mean_pool, global_max_pool
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import (roc_auc_score, accuracy_score, f1_score,
                             matthews_corrcoef, confusion_matrix)
from rdkit import Chem
import os, time, warnings
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, numbers)
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

# ─── Sabitler ─────────────────────────────────────────
TASKS = ['NR-AR','NR-AR-LBD','NR-AhR','NR-Aromatase','NR-ER','NR-ER-LBD',
         'NR-PPAR-gamma','SR-ARE','SR-ATAD5','SR-HSE','SR-MMP','SR-p53']
N_DESC  = 200
MAX_LEN = 100
EPOCHS  = 25
BATCH   = 64
LR      = 1e-3
DEVICE  = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

os.makedirs("results", exist_ok=True)
os.makedirs("models",  exist_ok=True)

# ─── Dataset sınıfları ────────────────────────────────
class TabDS(Dataset):
    """Descriptor/token tabanlı veri seti."""
    def __init__(self, X, y, dtype=torch.float32):
        self.X = torch.tensor(X, dtype=dtype)
        self.y = torch.tensor(y, dtype=torch.float32)
    def __len__(self): return len(self.y)
    def __getitem__(self, i): return self.X[i], self.y[i]

# ─── Moleküler graf oluşturma ─────────────────────────
def mol_to_graph(smiles, label):
    """
    SMILES → PyTorch Geometric Graf
    
    Her atom bir düğüm (node), her bağ bir kenar (edge).
    Atom özellikleri (9 adet):
      - Atom numarası / 118  (hangi element)
      - Bağ derecesi / 10    (kaç komşu)
      - Aromatiklik           (benzen halkasında mı)
      - H sayısı / 4          (kaç hidrojen)
      - Kütle / 200           (atom kütlesi)
      - Halka üyesi          (halkada mı)
      - Değerlik / 8          (toplam bağ kapasitesi)
      - Formal yük / 5        (elektrik yükü)
      - Radikal elektron / 4  (eşleşmemiş elektron)
    """
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        return None

    feats = []
    for atom in mol.GetAtoms():
        feats.append([
            atom.GetAtomicNum() / 118.0,
            atom.GetDegree() / 10.0,
            int(atom.GetIsAromatic()),
            atom.GetTotalNumHs() / 4.0,
            atom.GetMass() / 200.0,
            int(atom.IsInRing()),
            atom.GetTotalValence() / 8.0,
            atom.GetFormalCharge() / 5.0,
            atom.GetNumRadicalElectrons() / 4.0,
        ])
    x = torch.tensor(feats, dtype=torch.float)

    edge_list = []
    for bond in mol.GetBonds():
        i, j = bond.GetBeginAtomIdx(), bond.GetEndAtomIdx()
        edge_list += [[i, j], [j, i]]  # yönlü çift kenar

    if edge_list:
        edge_index = torch.tensor(edge_list, dtype=torch.long).t().contiguous()
    else:
        edge_index = torch.zeros((2, 0), dtype=torch.long)

    y = torch.tensor([float(label)], dtype=torch.float)
    return Data(x=x, edge_index=edge_index, y=y)


# ══════════════════════════════════════════════════════
#  MODEL 1: MLP (Multilayer Perceptron)
# ══════════════════════════════════════════════════════
class MLP(nn.Module):
    """
    Klasik tam bağlı sinir ağı.
    Girdi: 200 descriptor → Çıktı: toksisite olasılığı
    
    Katmanlar:
      Linear(200→512) → BatchNorm → ReLU → Dropout(0.3)
      Linear(512→256) → BatchNorm → ReLU → Dropout(0.3)
      Linear(256→128) → ReLU → Dropout(0.2)
      Linear(128→1)   → Sigmoid (BCELoss ile)
    """
    def __init__(self, in_dim=N_DESC):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(in_dim, 512),
            nn.BatchNorm1d(512),
            nn.ReLU(),
            nn.Dropout(0.3),
            nn.Linear(512, 256),
            nn.BatchNorm1d(256),
            nn.ReLU(),
            nn.Dropout(0.3),
            nn.Linear(256, 128),
            nn.ReLU(),
            nn.Dropout(0.2),
            nn.Linear(128, 1)
        )

    def forward(self, x):
        return self.net(x).squeeze(-1)


# ══════════════════════════════════════════════════════
#  MODEL 2: 1D-CNN (Bir Boyutlu Konvolüsyon)
# ══════════════════════════════════════════════════════
class CNN1D(nn.Module):
    """
    200 descriptor'ı 1D zaman serisi gibi işler.
    Konvolüsyon katmanları lokal descriptor örüntülerini yakalar.
    
    Örneğin: [MolLogP, TPSA, NumRings, ...] sırasındaki
    komşu descriptor'lar arasındaki ilişkileri öğrenir.
    
    Katmanlar:
      Conv1d(1→64, kernel=7)  → BN → ReLU → MaxPool(2)   → 100
      Conv1d(64→128, kernel=5) → BN → ReLU → MaxPool(2)  →  50
      Conv1d(128→256, kernel=3)→ BN → ReLU → AvgPool(1)  →   1
      Linear(256→128) → ReLU → Dropout(0.3)
      Linear(128→1)
    """
    def __init__(self):
        super().__init__()
        self.conv = nn.Sequential(
            nn.Conv1d(1, 64, kernel_size=7, padding=3),
            nn.BatchNorm1d(64),
            nn.ReLU(),
            nn.MaxPool1d(2),
            nn.Conv1d(64, 128, kernel_size=5, padding=2),
            nn.BatchNorm1d(128),
            nn.ReLU(),
            nn.MaxPool1d(2),
            nn.Conv1d(128, 256, kernel_size=3, padding=1),
            nn.BatchNorm1d(256),
            nn.ReLU(),
            nn.AdaptiveAvgPool1d(1)
        )
        self.fc = nn.Sequential(
            nn.Flatten(),
            nn.Linear(256, 128),
            nn.ReLU(),
            nn.Dropout(0.3),
            nn.Linear(128, 1)
        )

    def forward(self, x):
        return self.fc(self.conv(x)).squeeze(-1)


# ══════════════════════════════════════════════════════
#  MODEL 3: GNN — Graph Attention Network
# ══════════════════════════════════════════════════════
class GNN(nn.Module):
    """
    Moleküler grafu doğrudan işler — descriptor'a gerek yok!
    
    Her atom (düğüm) komşularından mesaj alır (message passing).
    GAT (Graph Attention): hangi komşunun daha önemli olduğunu öğrenir.
    
    Katmanlar:
      GATConv(9→128, 4 head) → LayerNorm → ELU    ← mesaj geçişi 1
      GATConv(128→128, 4 head)→ LayerNorm → ELU   ← mesaj geçişi 2
      GATConv(128→128, 2 head)→ LayerNorm → ELU   ← mesaj geçişi 3
      GlobalMeanPool + GlobalMaxPool → concat(256)  ← molekül temsili
      Linear(256→64) → ReLU → Linear(64→1)
    """
    def __init__(self):
        super().__init__()
        self.conv1 = GATConv(9,   128, heads=4, concat=False, dropout=0.2)
        self.conv2 = GATConv(128, 128, heads=4, concat=False, dropout=0.2)
        self.conv3 = GATConv(128, 128, heads=2, concat=False, dropout=0.2)
        self.ln1 = nn.LayerNorm(128)
        self.ln2 = nn.LayerNorm(128)
        self.ln3 = nn.LayerNorm(128)
        self.fc = nn.Sequential(
            nn.Linear(256, 256), nn.ReLU(), nn.Dropout(0.3),
            nn.Linear(256, 64),  nn.ReLU(),
            nn.Linear(64, 1)
        )

    def forward(self, data):
        x, ei, b = data.x, data.edge_index, data.batch
        x = F.elu(self.ln1(self.conv1(x, ei)))
        x = F.elu(self.ln2(self.conv2(x, ei)))
        x = F.elu(self.ln3(self.conv3(x, ei)))
        # Tüm atom temsillerini tek molekül vektörüne indir
        h = torch.cat([global_mean_pool(x, b),
                       global_max_pool(x, b)], dim=-1)
        return self.fc(h).squeeze(-1)


# ══════════════════════════════════════════════════════
#  MODEL 4: SMILES Transformer
# ══════════════════════════════════════════════════════
class SMILESTransformer(nn.Module):
    """
    SMILES string'ini doğrudan karakter dizisi olarak işler.
    NLP'deki BERT gibi çalışır ama kimyasal dil için.
    
    "CC(=O)Nc1ccc(O)cc1"  →  ['C','C','(','=','O',')','N','c','1',...]
    
    Katmanlar:
      Embedding(vocab→64)    ← karakter temsilleri
      PositionalEmbedding    ← sıra bilgisi
      TransformerEncoder × 2 ← self-attention ile bağlam
      LayerNorm → Linear(64→32) → GELU → Dropout(0.2) → Linear(32→1)
      
    CLS token (ilk token) molekül düzeyinde sınıflandırma yapar.
    """
    def __init__(self, vocab_size, d_model=64, nhead=4, num_layers=2):
        super().__init__()
        self.token_emb = nn.Embedding(vocab_size, d_model, padding_idx=0)
        self.pos_emb   = nn.Embedding(MAX_LEN, d_model)
        enc_layer = nn.TransformerEncoderLayer(
            d_model=d_model, nhead=nhead,
            dim_feedforward=256, dropout=0.1,
            batch_first=True, norm_first=True
        )
        self.transformer = nn.TransformerEncoder(enc_layer, num_layers=num_layers)
        self.classifier = nn.Sequential(
            nn.LayerNorm(d_model),
            nn.Linear(d_model, 32),
            nn.GELU(),
            nn.Dropout(0.2),
            nn.Linear(32, 1)
        )

    def forward(self, x):
        pos = torch.arange(x.size(1), device=x.device).unsqueeze(0)
        h   = self.token_emb(x) + self.pos_emb(pos)
        pad_mask = (x == 0)  # padding maskesi
        h   = self.transformer(h, src_key_padding_mask=pad_mask)
        return self.classifier(h[:, 0, :]).squeeze(-1)  # CLS token


# ─── Metrik hesaplama ─────────────────────────────────
def compute_metrics(y_true, y_prob):
    """ACC, AUC, F1, MCC, Sensitivity, Specificity hesapla."""
    y_pred = (y_prob > 0.5).astype(int)
    cm = confusion_matrix(y_true, y_pred, labels=[0, 1])
    tn, fp, fn, tp = cm.ravel() if cm.size == 4 else (0, 0, 0, 0)
    return {
        'ACC': round(accuracy_score(y_true, y_pred) * 100, 2),
        'AUC': round(roc_auc_score(y_true, y_prob), 4),
        'F1':  round(f1_score(y_true, y_pred, zero_division=0), 4),
        'MCC': round(matthews_corrcoef(y_true, y_pred), 4),
        'SE':  round(tp / (tp + fn) if (tp + fn) > 0 else 0, 4),  # Sensitivity
        'SP':  round(tn / (tn + fp) if (tn + fp) > 0 else 0, 4),  # Specificity
    }


def evaluate(model, loader, is_gnn=False):
    """Model değerlendirme — test seti metrikleri."""
    model.eval()
    probs, labels = [], []
    with torch.no_grad():
        if is_gnn:
            for batch in loader:
                batch = batch.to(DEVICE)
                p = torch.sigmoid(model(batch)).cpu().numpy().flatten()
                probs.extend(p)
                labels.extend(batch.y.cpu().numpy().flatten())
        else:
            for X, y in loader:
                X = X.to(DEVICE)
                p = torch.sigmoid(model(X)).cpu().numpy().flatten()
                probs.extend(p)
                labels.extend(y.numpy())
    y_true = np.array(labels)
    y_prob = np.clip(np.array(probs), 1e-7, 1 - 1e-7)
    return compute_metrics(y_true, y_prob)


# ─── Eğitim döngüsü ───────────────────────────────────
def train_model(model, train_dl, val_dl, optimizer, scheduler,
                criterion, epochs, is_gnn=False, model_name=""):
    """Genel eğitim döngüsü — early stopping ile."""
    best_auc   = 0
    best_state = None

    for epoch in range(epochs):
        model.train()
        if is_gnn:
            for batch in train_dl:
                batch = batch.to(DEVICE)
                optimizer.zero_grad()
                out = model(batch).view(-1)
                lbl = batch.y.view(-1).float()[:len(out)]
                # DÜZELTİLDİ: criterion (pos_weight dahil) kullan
                # Eski: F.binary_cross_entropy_with_logits(out, lbl)
                # Yeni: criterion(out, lbl)  ← pos_weight=5.3 artık GNN'e de uygulanır
                loss = criterion(out, lbl)
                loss.backward()
                optimizer.step()
        else:
            for X, y in train_dl:
                X, y = X.to(DEVICE), y.to(DEVICE)
                optimizer.zero_grad()
                loss = criterion(model(X), y)
                loss.backward()
                if model_name == "Transformer":
                    torch.nn.utils.clip_grad_norm_(model.parameters(), 1.0)
                optimizer.step()
        scheduler.step()

        # Validation AUC'ye göre en iyi modeli kaydet
        val_m = evaluate(model, val_dl, is_gnn)
        if val_m['AUC'] > best_auc:
            best_auc   = val_m['AUC']
            best_state = {k: v.clone() for k, v in model.state_dict().items()}

        if (epoch + 1) % 5 == 0:
            print(f"    Epoch {epoch+1:2d}/{epochs} | Val AUC: {val_m['AUC']:.4f}")

    model.load_state_dict(best_state)
    return model


# ─── Tek görev için 4 model ───────────────────────────
def run_task(task, df_full, X_desc_all, X_tok_all, vocab_size):
    """Belirtilen görev için 4 modeli eğit ve karşılaştır."""
    print(f"\n{'='*60}")
    print(f"  GÖREV: {task}")
    print(f"{'='*60}")

    # Göreve ait veriyi hazırla
    df_t = df_full[['smiles', task]].dropna()
    df_t = df_t[df_t[task].isin([0.0, 1.0])].drop_duplicates('smiles').reset_index(drop=True)
    df_t = df_t[df_t['smiles'].apply(lambda s: Chem.MolFromSmiles(s) is not None)]
    df_t = df_t.reset_index(drop=True)

    smiles = df_t['smiles'].tolist()
    labels = df_t[task].values
    all_smiles_list = df_full['smiles'].tolist()
    fi = np.array([all_smiles_list.index(s) if s in all_smiles_list else 0 for s in smiles])

    # Train / Val / Test böl (70 / 15 / 15)
    idx = np.arange(len(df_t))
    idx_tv, idx_te = train_test_split(idx, test_size=0.15, stratify=labels, random_state=42)
    idx_tr, idx_vl = train_test_split(idx_tv, test_size=0.176, stratify=labels[idx_tv], random_state=42)
    ytr, yvl, yte = labels[idx_tr], labels[idx_vl], labels[idx_te]

    # Sınıf dengesizliği için pos_weight
    pw = torch.tensor([(ytr == 0).sum() / max((ytr == 1).sum(), 1)], dtype=torch.float32).to(DEVICE)
    crit = nn.BCEWithLogitsLoss(pos_weight=pw)

    print(f"  n={len(df_t)} | train={len(idx_tr)} | val={len(idx_vl)} | test={len(idx_te)}")
    print(f"  Pozitif: {ytr.sum()} ({ytr.mean()*100:.1f}%) | pos_weight: {pw.item():.1f}")

    task_results = {}

    # ── MODEL 1: MLP ──────────────────────────────────
    print("\n  [1/4] MLP")
    sc   = StandardScaler()
    Xtr  = np.nan_to_num(sc.fit_transform(X_desc_all[fi][idx_tr]))
    Xvl  = np.nan_to_num(sc.transform(X_desc_all[fi][idx_vl]))
    Xte  = np.nan_to_num(sc.transform(X_desc_all[fi][idx_te]))

    tr1  = DataLoader(TabDS(Xtr, ytr), BATCH, shuffle=True, drop_last=True)
    vl1  = DataLoader(TabDS(Xvl, yvl), BATCH)
    te1  = DataLoader(TabDS(Xte, yte), BATCH)

    mlp  = MLP().to(DEVICE)
    opt1 = torch.optim.Adam(mlp.parameters(), lr=LR, weight_decay=1e-4)
    sch1 = torch.optim.lr_scheduler.CosineAnnealingLR(opt1, T_max=EPOCHS)
    t0   = time.time()
    mlp  = train_model(mlp, tr1, vl1, opt1, sch1, crit, EPOCHS, model_name="MLP")
    task_results['MLP'] = {**evaluate(mlp, te1), 'time_s': round(time.time()-t0, 1)}
    torch.save(mlp.state_dict(), f"models/mlp_{task}.pt")
    m = task_results['MLP']
    print(f"    → ACC:{m['ACC']}% AUC:{m['AUC']} F1:{m['F1']} MCC:{m['MCC']}")

    # ── MODEL 2: 1D-CNN ───────────────────────────────
    print("\n  [2/4] 1D-CNN")
    class CNN_DS(Dataset):
        def __init__(self, X, y):
            self.X = torch.tensor(X, dtype=torch.float32).unsqueeze(1)
            self.y = torch.tensor(y, dtype=torch.float32)
        def __len__(self): return len(self.y)
        def __getitem__(self, i): return self.X[i], self.y[i]

    tr2  = DataLoader(CNN_DS(Xtr, ytr), BATCH, shuffle=True, drop_last=True)
    vl2  = DataLoader(CNN_DS(Xvl, yvl), BATCH)
    te2  = DataLoader(CNN_DS(Xte, yte), BATCH)

    cnn  = CNN1D().to(DEVICE)
    opt2 = torch.optim.Adam(cnn.parameters(), lr=LR, weight_decay=1e-4)
    sch2 = torch.optim.lr_scheduler.CosineAnnealingLR(opt2, T_max=EPOCHS)
    t0   = time.time()
    cnn  = train_model(cnn, tr2, vl2, opt2, sch2, crit, EPOCHS, model_name="CNN")
    task_results['1D-CNN'] = {**evaluate(cnn, te2), 'time_s': round(time.time()-t0, 1)}
    torch.save(cnn.state_dict(), f"models/cnn_{task}.pt")
    m = task_results['1D-CNN']
    print(f"    → ACC:{m['ACC']}% AUC:{m['AUC']} F1:{m['F1']} MCC:{m['MCC']}")

    # ── MODEL 3: GNN ──────────────────────────────────
    print("\n  [3/4] GNN (GAT)")
    gtr_list = [mol_to_graph(smiles[i], ytr[j]) for j, i in enumerate(idx_tr)]
    gvl_list = [mol_to_graph(smiles[i], yvl[j]) for j, i in enumerate(idx_vl)]
    gte_list = [mol_to_graph(smiles[i], yte[j]) for j, i in enumerate(idx_te)]
    gtr_list = [g for g in gtr_list if g]
    gvl_list = [g for g in gvl_list if g]
    gte_list = [g for g in gte_list if g]

    gnn  = GNN().to(DEVICE)
    opt3 = torch.optim.Adam(gnn.parameters(), lr=LR, weight_decay=1e-4)
    sch3 = torch.optim.lr_scheduler.CosineAnnealingLR(opt3, T_max=EPOCHS)

    tr3  = GeoLoader(gtr_list, BATCH, shuffle=True, drop_last=True)
    vl3  = GeoLoader(gvl_list, BATCH)
    te3  = GeoLoader(gte_list, BATCH)
    t0   = time.time()
    gnn  = train_model(gnn, tr3, vl3, opt3, sch3, crit, EPOCHS, is_gnn=True)
    task_results['GNN'] = {**evaluate(gnn, te3, is_gnn=True), 'time_s': round(time.time()-t0, 1)}
    torch.save(gnn.state_dict(), f"models/gnn_{task}.pt")
    m = task_results['GNN']
    print(f"    → ACC:{m['ACC']}% AUC:{m['AUC']} F1:{m['F1']} MCC:{m['MCC']}")

    # ── MODEL 4: SMILES Transformer ───────────────────
    print("\n  [4/4] SMILES Transformer")
    Xt   = X_tok_all[fi]
    tr4  = DataLoader(TabDS(Xt[idx_tr], ytr, dtype=torch.long), BATCH, shuffle=True)
    vl4  = DataLoader(TabDS(Xt[idx_vl], yvl, dtype=torch.long), BATCH)
    te4  = DataLoader(TabDS(Xt[idx_te], yte, dtype=torch.long), BATCH)

    tfm  = SMILESTransformer(vocab_size).to(DEVICE)
    opt4 = torch.optim.AdamW(tfm.parameters(), lr=5e-4, weight_decay=1e-2)
    sch4 = torch.optim.lr_scheduler.CosineAnnealingLR(opt4, T_max=EPOCHS)
    t0   = time.time()
    tfm  = train_model(tfm, tr4, vl4, opt4, sch4, crit, EPOCHS, model_name="Transformer")
    task_results['Transformer'] = {**evaluate(tfm, te4), 'time_s': round(time.time()-t0, 1)}
    torch.save(tfm.state_dict(), f"models/tfm_{task}.pt")
    m = task_results['Transformer']
    print(f"    → ACC:{m['ACC']}% AUC:{m['AUC']} F1:{m['F1']} MCC:{m['MCC']}")

    return task_results


# ─── Excel çıktısı ────────────────────────────────────
def save_excel(df_res, all_results, path):
    """
    Sonuçları biçimlendirilmiş Excel dosyasına yazar.
    Sayfa 1: Tüm metrikler (ham tablo)
    Sayfa 2: AUC karşılaştırma özeti
    """
    # ── Renk ve stil sabitleri ──────────────────────────
    HEADER_FILL  = PatternFill("solid", start_color="1F3864")  # koyu lacivert
    HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    TASK_FILL    = PatternFill("solid", start_color="D6E4F0")   # açık mavi
    TASK_FONT    = Font(name="Arial", bold=True, size=10)
    BEST_FILL    = PatternFill("solid", start_color="E2EFDA")   # açık yeşil
    NORMAL_FONT  = Font(name="Arial", size=10)
    CENTER       = Alignment(horizontal="center", vertical="center")
    LEFT         = Alignment(horizontal="left",   vertical="center")
    THIN         = Side(style="thin", color="BFBFBF")
    BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    MODEL_COLORS = {
        "MLP":         "EAF4FB",
        "1D-CNN":      "F3EAF8",
        "GNN":         "FEF0E7",
        "Transformer": "EAFAF1",
    }

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════
    #  SAYFA 1 — Tüm metrikler
    # ══════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Tüm Metrikler"
    ws1.freeze_panes = "C2"   # ilk 2 sütun ve başlık satırı sabit

    cols = ["Task", "Model", "ACC", "AUC", "F1", "MCC", "SE", "SP", "time_s"]
    col_widths = [16, 14, 9, 9, 9, 9, 9, 9, 10]
    col_labels = ["Görev", "Model", "ACC (%)", "AUC", "F1", "MCC", "SE", "SP", "Süre (s)"]

    # Başlık satırı
    for ci, (label, width) in enumerate(zip(col_labels, col_widths), 1):
        cell = ws1.cell(row=1, column=ci, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws1.column_dimensions[get_column_letter(ci)].width = width

    # Satır yüksekliği
    ws1.row_dimensions[1].height = 22

    # Veri satırları
    for ri, row in enumerate(df_res.itertuples(index=False), 2):
        task  = getattr(row, "Task")
        model = getattr(row, "Model")
        model_fill = PatternFill("solid", start_color=MODEL_COLORS.get(model, "FFFFFF"))

        for ci, col in enumerate(cols, 1):
            val  = getattr(row, col, "")
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.border    = BORDER
            cell.alignment = CENTER if ci > 2 else LEFT

            if ci == 1:   # Görev sütunu
                cell.font = TASK_FONT
                cell.fill = TASK_FILL
            elif ci == 2:  # Model sütunu
                cell.font = Font(name="Arial", size=10,
                                 color="FFFFFF" if model == "MLP" else "1F3864",
                                 bold=True)
                cell.fill = PatternFill("solid",
                    start_color="2E75B6" if model == "MLP" else
                                "7030A0" if model == "1D-CNN" else
                                "C55A11" if model == "GNN" else "375623")
            else:
                cell.font = NORMAL_FONT
                cell.fill = model_fill

            # AUC sütununu (ci=4) kalın yaz
            if ci == 4 and isinstance(val, float):
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.number_format = "0.0000"
            elif ci == 3 and isinstance(val, float):
                cell.number_format = "0.00"
            elif ci in (5, 6, 7, 8) and isinstance(val, float):
                cell.number_format = "0.0000"

        ws1.row_dimensions[ri].height = 18

    # ══════════════════════════════════════════════════
    #  SAYFA 2 — AUC Özet
    # ══════════════════════════════════════════════════
    ws2 = wb.create_sheet("AUC Özeti")
    ws2.freeze_panes = "B2"

    mcols  = ["MLP", "1D-CNN", "GNN", "Transformer"]
    tasks  = df_res["Task"].unique().tolist()

    # Başlık satırı
    ws2.cell(row=1, column=1, value="Görev").font      = HEADER_FONT
    ws2.cell(row=1, column=1).fill      = HEADER_FILL
    ws2.cell(row=1, column=1).alignment = CENTER
    ws2.cell(row=1, column=1).border    = BORDER
    ws2.column_dimensions["A"].width    = 18

    for ci, mc in enumerate(mcols, 2):
        cell = ws2.cell(row=1, column=ci, value=mc)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws2.column_dimensions[get_column_letter(ci)].width = 13

    ws2.column_dimensions[get_column_letter(len(mcols)+2)].width = 13
    ws2.cell(row=1, column=len(mcols)+2, value="En İyi Model").font      = HEADER_FONT
    ws2.cell(row=1, column=len(mcols)+2).fill      = HEADER_FILL
    ws2.cell(row=1, column=len(mcols)+2).alignment = CENTER
    ws2.cell(row=1, column=len(mcols)+2).border    = BORDER
    ws2.row_dimensions[1].height = 22

    # Görev satırları
    for ri, task in enumerate(tasks, 2):
        task_models = all_results.get(task, {})
        aucs = [task_models.get(mc, {}).get("AUC", float("nan")) for mc in mcols]
        valid = [a for a in aucs if not (a != a)]  # nan filtrele
        best_auc = max(valid) if valid else float("nan")

        # Görev adı
        tc = ws2.cell(row=ri, column=1, value=task)
        tc.font = TASK_FONT; tc.fill = TASK_FILL
        tc.alignment = LEFT; tc.border = BORDER

        for ci, (mc, auc) in enumerate(zip(mcols, aucs), 2):
            cell = ws2.cell(row=ri, column=ci,
                            value=round(auc, 4) if auc == auc else "N/A")
            cell.alignment    = CENTER
            cell.border       = BORDER
            cell.number_format = "0.0000"
            if auc == auc:  # nan değil
                cell.font = Font(name="Arial", size=10,
                                 bold=(auc == best_auc))
                cell.fill = (BEST_FILL if auc == best_auc
                             else PatternFill("solid",
                                  start_color=MODEL_COLORS.get(mc, "FFFFFF")))
            else:
                cell.font = NORMAL_FONT

        # En iyi model sütunu
        if valid:
            best_name = mcols[aucs.index(best_auc)]
        else:
            best_name = "N/A"
        bc = ws2.cell(row=ri, column=len(mcols)+2, value=best_name)
        bc.font = Font(name="Arial", size=10, bold=True, color="375623")
        bc.fill = BEST_FILL; bc.alignment = CENTER; bc.border = BORDER
        ws2.row_dimensions[ri].height = 18

    # Ortalama satırı (birden fazla görev varsa)
    if len(tasks) > 1:
        avg_row = len(tasks) + 2
        ac = ws2.cell(row=avg_row, column=1, value="Ortalama")
        ac.font = Font(name="Arial", bold=True, size=10)
        ac.fill = PatternFill("solid", start_color="F2F2F2")
        ac.alignment = LEFT; ac.border = BORDER

        means = []
        for ci, mc in enumerate(mcols, 2):
            vals = [all_results[t].get(mc, {}).get("AUC", float("nan"))
                    for t in tasks]
            vals = [v for v in vals if v == v]
            mean_val = sum(vals)/len(vals) if vals else float("nan")
            means.append(mean_val)
            cell = ws2.cell(row=avg_row, column=ci,
                            value=round(mean_val, 4) if mean_val == mean_val else "N/A")
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.fill = PatternFill("solid", start_color="F2F2F2")
            cell.alignment = CENTER; cell.border = BORDER
            cell.number_format = "0.0000"

        overall_best = mcols[means.index(max(m for m in means if m == m))]
        bc2 = ws2.cell(row=avg_row, column=len(mcols)+2, value=f"⭐ {overall_best}")
        bc2.font = Font(name="Arial", bold=True, size=10, color="375623")
        bc2.fill = BEST_FILL; bc2.alignment = CENTER; bc2.border = BORDER
        ws2.row_dimensions[avg_row].height = 20

    wb.save(path)
    print(f"  Excel dosyası oluşturuldu: {path}")


# ─── ANA FONKSİYON ────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--task', default='SR-MMP',
                        help='Görev adı veya "all" (örn: SR-MMP, NR-AhR)')
    args = parser.parse_args()

    # Veriyi yükle
    print("Veri yükleniyor...")
    df_full = pd.read_csv("data/tox21.csv")
    X_desc  = np.load("features/X_descriptors.npy")
    X_tok   = np.load("features/X_tokens.npy")
    vocab   = ['<PAD>', '<UNK>'] + sorted(set(''.join(df_full['smiles'].tolist())))
    print(f"  {len(df_full)} bileşik | Descriptor: {X_desc.shape} | Token: {X_tok.shape}")

    # Görev(ler)
    selected_tasks = TASKS if args.task == 'all' else [args.task]
    all_results = {}

    for task in selected_tasks:
        task_res = run_task(task, df_full, X_desc, X_tok, len(vocab))
        all_results[task] = task_res

    # Sonuçları Excel'e kaydet
    rows = []
    for task, models in all_results.items():
        for mname, mvals in models.items():
            rows.append({'Task': task, 'Model': mname, **mvals})
    df_res = pd.DataFrame(rows)
    save_excel(df_res, all_results, "results/model_comparison.xlsx")

    # Özet tablo
    print(f"\n{'='*70}")
    print("  ÖZET — AUC Karşılaştırması")
    print(f"{'='*70}")
    mcols = ['MLP', '1D-CNN', 'GNN', 'Transformer']
    print(f"  {'Görev':<20}", end="")
    for c in mcols: print(f"  {c:>12}", end="")
    print("  En İyi")
    print("  " + "─"*65)

    for task, models in all_results.items():
        aucs = [models.get(m, {}).get('AUC', np.nan) for m in mcols]
        best = mcols[np.nanargmax(aucs)]
        print(f"  {task:<20}", end="")
        for a in aucs:
            print(f"  {a:>12.4f}" if not np.isnan(a) else f"  {'N/A':>12}", end="")
        print(f"  ← {best}")

    if len(selected_tasks) > 1:
        means = []
        for mc in mcols:
            vals = [all_results[t].get(mc, {}).get('AUC', np.nan) for t in selected_tasks]
            means.append(np.nanmean(vals))
        best_model = mcols[np.nanargmax(means)]
        print("  " + "─"*65)
        print(f"  {'Ortalama':<20}", end="")
        for v in means: print(f"  {v:>12.4f}", end="")
        print(f"\n\n  ⭐ En iyi model: {best_model} (Ort. AUC={max(means):.4f})")

    print(f"\n✅ Sonuçlar kaydedildi: results/model_comparison.xlsx")
    print(f"✅ Model ağırlıkları kaydedildi: models/")


if __name__ == "__main__":
    main()